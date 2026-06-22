import csv
import io
import secrets
from typing import Any, Optional
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException, status, BackgroundTasks

from app.repositories.employee_repository import employee_repository
from app.models.employee import Employee, EmployeeStatus
from app.schemas.employee import EmployeeCreate, EmployeeRegister, EmployeeUpdate
from app.repositories.company_repository import company_repository
from app.repositories.user_repository import user_repository
from app.models.user import User, UserRole
from app.core.security.security import hash_password
from app.services.notification_service import notification_service
from app.services.audit_service import audit_log_service


class EmployeeService:
    @staticmethod
    async def create_employee(
        db: AsyncSession, company_id: int, obj_in: EmployeeCreate, actor_id: Optional[int] = None,
        background_tasks: Optional[BackgroundTasks] = None
    ) -> Employee:
        company = await company_repository.get(db, company_id)
        if not company:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")

        if company.current_employee_count >= company.employee_limit:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Company employee limit reached. Upgrade subscription plan.",
            )

        existing_user = await user_repository.get_by_email(db, obj_in.email)
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Email {obj_in.email} is already registered.",
            )

        temp_pass = secrets.token_urlsafe(10)[:12]
        hashed_pass = hash_password(temp_pass)

        user = User(
            email=obj_in.email,
            password_hash=hashed_pass,
            role=UserRole.EMPLOYEE,
            company_id=company_id,
            is_active=True,
            is_verified=False,
        )
        db.add(user)
        await db.commit()
        await db.refresh(user)

        employee = Employee(
            id=user.id,
            company_id=company_id,
            department_id=obj_in.department_id,
            first_name=obj_in.first_name,
            last_name=obj_in.last_name,
            email=obj_in.email,
            status=EmployeeStatus.INVITED,
        )
        db.add(employee)
        await db.commit()
        await db.refresh(employee)

        await company_repository.update_employee_count(db, company_id=company_id, count_change=1)

        await audit_log_service.log_action(
            db=db,
            actor_id=actor_id,
            action="CREATE_EMPLOYEE",
            entity_type="Employee",
            entity_id=employee.id,
            new_value={"email": obj_in.email, "company_id": company_id},
        )

        from app.tasks.queue import enqueue_task
        from app.tasks.worker import send_onboarding_email_job
        if background_tasks:
            await enqueue_task(
                background_tasks,
                send_onboarding_email_job,
                email=obj_in.email,
                first_name=obj_in.first_name,
                temp_pass=temp_pass,
            )
        else:
            await notification_service.send_onboarding_email(
                email=obj_in.email,
                first_name=obj_in.first_name,
                temp_pass=temp_pass,
            )

        return employee

    @staticmethod
    async def bulk_upload_employees(
        db: AsyncSession, company_id: int, file_content: bytes, filename: str,
        actor_id: Optional[int] = None, background_tasks: Optional[BackgroundTasks] = None
    ) -> dict[str, Any]:
        records = []
        errors = []

        if filename.endswith(".csv"):
            try:
                stream = io.StringIO(file_content.decode("utf-8"))
                reader = csv.DictReader(stream)
                for row in reader:
                    records.append(row)
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Failed to parse CSV file: {e}",
                )
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content))
                sheet = wb.active
                header = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(max_row=1))]

                if "email" not in header:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="XLSX must contain an 'email' column.",
                    )

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = dict(zip(header, row))
                    records.append(row_dict)
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Failed to parse XLSX file: {e}",
                )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format. Please upload CSV or XLSX.",
            )

        imported_count = 0
        failed_count = 0

        for idx, rec in enumerate(records):
            email = rec.get("email")
            name = rec.get("name")
            first_name = rec.get("first_name")
            last_name = rec.get("last_name")

            if not email or not isinstance(email, str) or not email.strip():
                errors.append(f"Row {idx+2}: Missing email address.")
                failed_count += 1
                continue

            email = email.strip()

            if name and not first_name:
                name_parts = str(name).strip().split(maxsplit=1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ""
            else:
                first_name = str(first_name or "").strip() or "Employee"
                last_name = str(last_name or "").strip() or ""

            emp_in = EmployeeCreate(
                first_name=first_name,
                last_name=last_name,
                email=email,
                department_id=None,
            )

            try:
                await EmployeeService.create_employee(
                    db=db, company_id=company_id, obj_in=emp_in,
                    actor_id=actor_id, background_tasks=background_tasks
                )
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {idx+2} ({email}): {str(e)}")
                failed_count += 1

        return {
            "total": len(records),
            "imported": imported_count,
            "failed": failed_count,
            "errors": errors,
        }

    @staticmethod
    async def register_employee(db: AsyncSession, employee_id: int, obj_in: EmployeeRegister) -> Employee:
        employee = await employee_repository.get(db, employee_id)
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")

        user = await user_repository.get(db, employee_id)
        if user:
            user.is_verified = True
            db.add(user)

        update_data = EmployeeUpdate(
            employee_code=obj_in.employee_code,
            phone=obj_in.phone,
            gender=obj_in.gender,
            date_of_birth=obj_in.date_of_birth,
            address=obj_in.address,
            joining_date=obj_in.joining_date,
            status=EmployeeStatus.REGISTERED,
        )

        updated_emp = await employee_repository.update(db, db_obj=employee, obj_in=update_data)
        updated_emp.registration_completed = True
        db.add(updated_emp)
        await db.commit()
        await db.refresh(updated_emp)

        await audit_log_service.log_action(
            db=db,
            actor_id=employee_id,
            action="EMPLOYEE_REGISTRATION",
            entity_type="Employee",
            entity_id=employee_id,
        )

        return updated_emp


employee_service = EmployeeService()